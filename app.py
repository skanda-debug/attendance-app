"""
Dynamic QR Code Attendance System - SERVER
Run this on the TEACHER's computer / projector screen.

UPGRADES:
  - Multi-threaded HTTP server
  - PermissionError fix
  - Branch & Section field
  - Late entry warning
  - Admin panel (view/delete)
  - Settings screen (change subject, timer, password from UI)
  - Absent list auto-generated in Excel
  - Monthly attendance summary sheet
  - ONE DEVICE ONE SUBMISSION: same phone/device cannot submit twice per subject session
"""

import qrcode
import threading
import time
import os
import socket
import secrets
import string
from datetime import datetime, date
from calendar import monthrange
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import parse_qs, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from PIL import Image, ImageTk
import io


# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
QR_ROTATE_SECONDS  = 30
SERVER_PORT        = 5050
EXCEL_FILENAME     = "attendance.xlsx"
SUBJECT            = "DTL"
LATE_AFTER_MINUTES = 10
CLASS_START_TIME   = None
ADMIN_PASSWORD     = "admin123"
ROLL_LIST          = []


# ─────────────────────────────────────────────
#  SHARED STATE
# ─────────────────────────────────────────────
state_lock        = threading.Lock()
current_token     = ""
token_expiry      = 0.0
attendance_log    = []
class_start_dt    = None
submitted_ips     = set()   # IPs that already submitted this session


def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def generate_token(length=12):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def rotate_token():
    global current_token, token_expiry
    while True:
        new_token = generate_token()
        with state_lock:
            current_token = new_token
            token_expiry  = time.time() + QR_ROTATE_SECONDS
        time.sleep(QR_ROTATE_SECONDS)


def is_late():
    if class_start_dt is None:
        return False
    return (datetime.now() - class_start_dt).total_seconds() / 60 > LATE_AFTER_MINUTES


# ─────────────────────────────────────────────
#  HTML TEMPLATES
# ─────────────────────────────────────────────
HTML_FORM = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Mark Attendance</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{
    font-family:'Segoe UI',sans-serif;
    background:linear-gradient(135deg,#1a1a2e 0%,#16213e 50%,#0f3460 100%);
    min-height:100vh; display:flex; align-items:center;
    justify-content:center; padding:20px;
  }}
  .card {{
    background:white; border-radius:16px; padding:36px 32px;
    width:100%; max-width:420px; box-shadow:0 20px 60px rgba(0,0,0,0.4);
  }}
  .logo {{ text-align:center; margin-bottom:20px; }}
  .logo h1 {{ font-size:22px; color:#1a1a2e; font-weight:700; }}
  .logo p  {{ font-size:13px; color:#888; margin-top:4px; }}
  .late-badge {{
    background:#fef3c7; color:#92400e; border:1px solid #fde68a;
    border-radius:8px; padding:8px 12px; font-size:13px;
    text-align:center; margin-bottom:16px; font-weight:600;
  }}
  label {{ display:block; font-size:13px; font-weight:600; color:#444; margin-bottom:6px; }}
  input, select {{
    width:100%; padding:11px 14px; border:1.5px solid #ddd;
    border-radius:8px; font-size:15px; margin-bottom:16px;
    transition:border 0.2s; background:white;
  }}
  input:focus, select:focus {{ outline:none; border-color:#0f3460; }}
  .row {{ display:flex; gap:12px; }}
  .row > div {{ flex:1; }}
  button {{
    width:100%; padding:13px; background:#0f3460;
    color:white; border:none; border-radius:8px;
    font-size:16px; font-weight:600; cursor:pointer;
  }}
  button:hover {{ background:#1a4a8a; }}
  .timer-bar {{ height:4px; background:#e5e7eb; border-radius:2px; overflow:hidden; margin-bottom:20px; }}
  .timer-fill {{
    height:100%; background:#0f3460;
    animation:drain {interval}s linear forwards;
  }}
  @keyframes drain {{ from{{width:100%}} to{{width:0%}} }}
</style>
</head>
<body>
<div class="card">
  <div class="logo">
    <h1>&#x1F4CB; Mark Attendance</h1>
    <p>{date} &nbsp;|&nbsp; {subject}</p>
  </div>
  <div class="timer-bar"><div class="timer-fill"></div></div>
  {late_banner}
  <form method="POST" action="/submit?token={token}">
    <label>USN / Roll Number</label>
    <input name="usn" type="text" placeholder="e.g. 1BM21CS001" required autofocus>
    <label>Full Name</label>
    <input name="name" type="text" placeholder="e.g. Rahul Sharma" required>
    <div class="row">
      <div>
        <label>Branch</label>
        <select name="branch" required>
          <option value="" disabled selected>Select</option>
          <option value="CSE">CSE</option>
          <option value="ETE">ETE</option>
          <option value="ISE">ISE</option>
          <option value="ECE">ECE</option>
          <option value="EEE">EEE</option>
          <option value="ME">ME</option>
          <option value="CV">CV</option>
          <option value="AI&ML">AI&amp;ML</option>
          <option value="Other">Other</option>
        </select>
      </div>
      <div>
        <label>Section</label>
        <select name="section" required>
          <option value="" disabled selected>Select</option>
          <option value="No Section">No Section</option>
          <option value="A">A</option>
          <option value="B">B</option>
          <option value="C">C</option>
          <option value="D">D</option>
          <option value="E">E</option>
        </select>
      </div>
    </div>
    <button type="submit">Submit Attendance &#x2713;</button>
  </form>
</div>
</body>
</html>"""

HTML_SUCCESS = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Attendance Marked</title>
<style>
  body {{ font-family:'Segoe UI',sans-serif;
    background:linear-gradient(135deg,#1a1a2e,#0f3460);
    display:flex;align-items:center;justify-content:center;min-height:100vh; }}
  .card {{ background:white;border-radius:16px;padding:40px 32px;text-align:center;max-width:400px; }}
  .icon {{ font-size:60px; margin-bottom:12px; }}
  h2 {{ font-size:22px; }}
  .ok {{ color:#16a34a; }} .late {{ color:#d97706; }}
  p {{ color:#555; margin-top:8px; font-size:14px; }}
  .badge {{ display:inline-block;margin-top:12px;padding:4px 12px;border-radius:20px;font-size:13px;font-weight:600; }}
  .badge-present {{ background:#d1fae5; color:#065f46; }}
  .badge-late    {{ background:#fef3c7; color:#92400e; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">{icon}</div>
  <h2 class="{status_class}">Attendance Marked!</h2>
  <p><strong>{usn}</strong> &mdash; {name}</p>
  <p>{branch} &nbsp;|&nbsp; Section: {section}</p>
  <p style="margin-top:8px;color:#888">{date} &nbsp; {subject}</p>
  <span class="badge {badge_class}">{status}</span>
</div>
</body>
</html>"""

HTML_EXPIRED = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><title>QR Expired</title>
<style>
  body {{ font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#7f1d1d,#991b1b);
    display:flex;align-items:center;justify-content:center;min-height:100vh; }}
  .card {{ background:white;border-radius:16px;padding:40px;text-align:center;max-width:360px; }}
  .icon {{ font-size:60px; }} h2 {{ color:#dc2626;margin-top:12px; }}
  p {{ color:#555;margin-top:8px;font-size:14px; }}
</style></head>
<body><div class="card">
  <div class="icon">&#x23F1;&#xFE0F;</div>
  <h2>QR Code Expired</h2>
  <p>Please scan the latest QR code shown on the screen.</p>
</div></body></html>"""

HTML_DUPLICATE = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><title>Already Marked</title>
<style>
  body {{ font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#1e3a5f,#0f3460);
    display:flex;align-items:center;justify-content:center;min-height:100vh; }}
  .card {{ background:white;border-radius:16px;padding:40px;text-align:center;max-width:360px; }}
  .icon {{ font-size:60px; }} h2 {{ color:#d97706;margin-top:12px; }}
  p {{ color:#555;margin-top:8px;font-size:14px; }}
</style></head>
<body><div class="card">
  <div class="icon">&#x26A0;&#xFE0F;</div>
  <h2>Already Marked</h2>
  <p><strong>{usn}</strong> already marked at <strong>{time}</strong> today.</p>
</div></body></html>"""

HTML_DEVICE_BLOCKED = """<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><title>Device Already Used</title>
<style>
  body {{ font-family:'Segoe UI',sans-serif;
    background:linear-gradient(135deg,#312e81,#4338ca);
    display:flex;align-items:center;justify-content:center;min-height:100vh; }}
  .card {{ background:white;border-radius:16px;padding:40px;text-align:center;max-width:380px; }}
  .icon {{ font-size:60px; }} h2 {{ color:#4338ca;margin-top:12px; }}
  p {{ color:#555;margin-top:8px;font-size:14px;line-height:1.6; }}
  .note {{ margin-top:16px;background:#ede9fe;color:#4338ca;padding:10px 14px;
           border-radius:8px;font-size:13px;font-weight:600; }}
</style></head>
<body><div class="card">
  <div class="icon">&#x1F4F5;</div>
  <h2>Device Already Used</h2>
  <p>Attendance has already been submitted from this device for <strong>{subject}</strong> today.</p>
  <p>Each device can only submit attendance once per subject session.</p>
  <div class="note">If this is a mistake, ask your teacher to reset device locks.</div>
</div></body></html>"""

HTML_ADMIN = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Admin Panel</title>
<style>
  * {{ box-sizing:border-box;margin:0;padding:0; }}
  body {{ font-family:'Segoe UI',sans-serif;background:#0f172a;color:#e2e8f0;padding:24px; }}
  h1 {{ font-size:20px;margin-bottom:4px;color:#38bdf8; }}
  .sub {{ font-size:13px;color:#94a3b8;margin-bottom:20px; }}
  table {{ width:100%;border-collapse:collapse;font-size:13px; }}
  th {{ background:#0f3460;padding:10px 12px;text-align:left;color:white; }}
  td {{ padding:9px 12px;border-bottom:1px solid #1e293b; }}
  tr:hover td {{ background:#1e293b; }}
  .present {{ color:#4ade80;font-weight:600; }}
  .late    {{ color:#fbbf24;font-weight:600; }}
  .absent  {{ color:#f87171;font-weight:600; }}
  .del-btn {{ background:#7f1d1d;color:white;border:none;padding:4px 10px;border-radius:6px;cursor:pointer;font-size:12px; }}
  .del-btn:hover {{ background:#991b1b; }}
  .summary {{ margin-top:16px;font-size:14px;color:#94a3b8;background:#1e293b;padding:12px 16px;border-radius:8px; }}
  .summary span {{ font-weight:700; }}
  .sp {{ color:#4ade80; }} .sl {{ color:#fbbf24; }} .sa {{ color:#f87171; }}
  .reset-btn {{
    display:inline-block;margin-top:16px;padding:10px 20px;
    background:#312e81;color:white;border:none;border-radius:8px;
    font-size:13px;font-weight:600;cursor:pointer;text-decoration:none;
  }}
  .reset-btn:hover {{ background:#3730a3; }}
  .device-count {{ font-size:13px;color:#94a3b8;margin-top:8px; }}
  .device-count span {{ color:#a78bfa;font-weight:700; }}
</style></head>
<body>
<h1>Admin Panel</h1>
<div class="sub">{subject} &nbsp;|&nbsp; {date}</div>
<table>
  <tr><th>#</th><th>USN</th><th>Name</th><th>Branch</th><th>Section</th><th>Time</th><th>Status</th><th>Action</th></tr>
  {rows}
</table>
<div class="summary">
  Total: <span class="sp">{total}</span> &nbsp;|&nbsp;
  Present: <span class="sp">{present}</span> &nbsp;|&nbsp;
  Late: <span class="sl">{late}</span> &nbsp;|&nbsp;
  Absent: <span class="sa">{absent}</span>
</div>
<div class="device-count">Devices locked this session: <span>{device_count}</span></div>
<a href="/admin/reset_devices?pwd={pwd}" class="reset-btn"
   onclick="return confirm('Reset all device locks? Students can submit again from any device.')">
  Reset Device Locks
</a>
</body></html>"""


# ─────────────────────────────────────────────
#  THREADED HTTP SERVER
# ─────────────────────────────────────────────
class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


class AttendanceHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        parsed = urlparse(self.path)

        # ── Admin panel ──────────────────────────────────────────────────
        if parsed.path == "/admin":
            pwd = parse_qs(parsed.query).get("pwd", [""])[0]
            if pwd != ADMIN_PASSWORD:
                self._send_html(401, "<h2 style='color:red;font-family:sans-serif;padding:20px'>Wrong password.</h2>")
                return
            self._send_html(200, self._build_admin_html())
            return

        # ── Delete entry ─────────────────────────────────────────────────
        if parsed.path == "/admin/delete":
            pwd = parse_qs(parsed.query).get("pwd", [""])[0]
            usn = parse_qs(parsed.query).get("usn", [""])[0]
            if pwd != ADMIN_PASSWORD:
                self._send_html(401, "<h2>Unauthorized</h2>")
                return
            today = datetime.now().strftime("%Y-%m-%d")
            with state_lock:
                before = len(attendance_log)
                attendance_log[:] = [r for r in attendance_log
                                     if not (r["USN"] == usn and r["Date"] == today)]
                removed = before - len(attendance_log)
            if removed:
                rebuild_excel_for_today()
            self._send_html(200, "<meta http-equiv='refresh' content='0;url=/admin?pwd=" + pwd + "'><p>Deleted</p>")
            return

        # ── Reset device locks ───────────────────────────────────────────
        if parsed.path == "/admin/reset_devices":
            pwd = parse_qs(parsed.query).get("pwd", [""])[0]
            if pwd != ADMIN_PASSWORD:
                self._send_html(401, "<h2>Unauthorized</h2>")
                return
            with state_lock:
                submitted_ips.clear()
            self._send_html(200, "<meta http-equiv='refresh' content='0;url=/admin?pwd=" + pwd + "'>"
                                 "<p>Device locks cleared.</p>")
            return

        # ── Main attendance form ─────────────────────────────────────────
        token_param = parse_qs(parsed.query).get("token", [""])[0]
        with state_lock:
            valid  = (token_param == current_token and time.time() < token_expiry)
            token  = current_token
            expiry = token_expiry

        if not valid:
            self._send_html(403, HTML_EXPIRED)
            return

        remaining   = max(0, int(expiry - time.time()))
        late_banner = '<div class="late-badge">You are marking attendance late!</div>' if is_late() else ""
        self._send_html(200, HTML_FORM.format(
            token=token, date=datetime.now().strftime("%d %B %Y"),
            subject=SUBJECT, interval=remaining, late_banner=late_banner
        ))

    def do_POST(self):
        parsed      = urlparse(self.path)
        token_param = parse_qs(parsed.query).get("token", [""])[0]

        with state_lock:
            valid = (token_param == current_token and time.time() < token_expiry)

        if not valid:
            self._send_html(403, HTML_EXPIRED)
            return

        # ── Device (IP) check ────────────────────────────────────────────
        client_ip = self.client_address[0]
        with state_lock:
            if client_ip in submitted_ips:
                self._send_html(200, HTML_DEVICE_BLOCKED.format(subject=SUBJECT))
                return

        length  = int(self.headers.get("Content-Length", 0))
        body    = self.rfile.read(length).decode()
        fields  = parse_qs(body)
        usn     = fields.get("usn",     [""])[0].strip().upper()
        name    = fields.get("name",    [""])[0].strip().title()
        branch  = fields.get("branch",  [""])[0].strip()
        section = fields.get("section", [""])[0].strip()
        today   = datetime.now().strftime("%Y-%m-%d")

        if not usn or not name or not branch or not section:
            self._send_html(400, "<h2 style='font-family:sans-serif;padding:20px'>Please fill all fields.</h2>")
            return

        with state_lock:
            # USN duplicate check
            dup = next((r for r in attendance_log if r["USN"] == usn and r["Date"] == today), None)
            if dup:
                self._send_html(200, HTML_DUPLICATE.format(usn=usn, time=dup["Time"]))
                return

            status = "Late" if is_late() else "Present"
            record = {
                "USN": usn, "Name": name, "Branch": branch, "Section": section,
                "Date": today, "Time": datetime.now().strftime("%H:%M:%S"),
                "Subject": SUBJECT, "Status": status
            }
            attendance_log.append(record)
            submitted_ips.add(client_ip)   # lock this device

        save_to_excel(record)
        self._send_html(200, HTML_SUCCESS.format(
            icon="&#x2705;" if status == "Present" else "&#x23F0;",
            status_class="ok" if status == "Present" else "late",
            badge_class="badge-present" if status == "Present" else "badge-late",
            status=status, usn=usn, name=name, branch=branch, section=section,
            date=datetime.now().strftime("%d %B %Y"), subject=SUBJECT
        ))

    def _build_admin_html(self):
        today = datetime.now().strftime("%Y-%m-%d")
        with state_lock:
            records      = [r for r in attendance_log if r["Date"] == today]
            device_count = len(submitted_ips)
        rows = ""
        for i, r in enumerate(records, 1):
            sc = "late" if r["Status"] == "Late" else ("absent" if r["Status"] == "Absent" else "present")
            del_btn = (f"<button class='del-btn' "
                       f"onclick=\"location.href='/admin/delete?pwd={ADMIN_PASSWORD}&usn={r['USN']}'\""
                       f">Delete</button>") if r["Status"] != "Absent" else "—"
            rows += (f"<tr><td>{i}</td><td>{r['USN']}</td><td>{r['Name']}</td>"
                     f"<td>{r['Branch']}</td><td>{r['Section']}</td><td>{r['Time']}</td>"
                     f"<td class='{sc}'>{r['Status']}</td><td>{del_btn}</td></tr>")
        present = sum(1 for r in records if r["Status"] == "Present")
        late    = sum(1 for r in records if r["Status"] == "Late")
        absent  = sum(1 for r in records if r["Status"] == "Absent")
        return HTML_ADMIN.format(
            subject=SUBJECT, date=datetime.now().strftime("%d %B %Y"),
            pwd=ADMIN_PASSWORD,
            rows=rows if rows else "<tr><td colspan='8' style='text-align:center;padding:20px;color:#64748b'>No records yet</td></tr>",
            total=len(records), present=present, late=late, absent=absent,
            device_count=device_count
        )

    def _send_html(self, code, html):
        data = html.encode()
        self.send_response(code)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


# ─────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────
HEADERS      = ["USN", "Name", "Branch", "Section", "Date", "Time", "Subject", "Status"]
COL_WIDTHS   = [16, 22, 10, 12, 14, 12, 14, 12]
HEADER_FILL  = PatternFill("solid", start_color="0F3460")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
PRESENT_FILL = PatternFill("solid", start_color="D1FAE5")
LATE_FILL    = PatternFill("solid", start_color="FEF3C7")
ABSENT_FILL  = PatternFill("solid", start_color="FEE2E2")
CELL_FONT    = Font(name="Arial", size=10)
THIN         = Side(style="thin", color="CCCCCC")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER       = Alignment(horizontal="center", vertical="center")


def _write_header(ws):
    ws.append(HEADERS)
    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h; cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22


def _load_or_create_wb(path):
    if os.path.exists(path):
        try:
            return openpyxl.load_workbook(path)
        except Exception:
            pass
    return openpyxl.Workbook()


def _style_row(ws, row_num):
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = CENTER
        val = cell.value
        if val == "Present":  cell.fill = PRESENT_FILL
        elif val == "Late":   cell.fill = LATE_FILL
        elif val == "Absent": cell.fill = ABSENT_FILL


def _write_summary(ws):
    for row in range(ws.max_row, 1, -1):
        v = ws.cell(row, 1).value
        if v in ("Total Present", "Total Late", "Total Absent", None):
            ws.delete_rows(row)
        else:
            break
    present = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[7] == "Present")
    late    = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[7] == "Late")
    absent  = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[7] == "Absent")
    sr = ws.max_row + 2
    ws.cell(sr,   1, "Total Present").font = Font(bold=True, name="Arial")
    ws.cell(sr,   2, present).font         = Font(bold=True, name="Arial", color="16A34A")
    ws.cell(sr+1, 1, "Total Late").font    = Font(bold=True, name="Arial")
    ws.cell(sr+1, 2, late).font            = Font(bold=True, name="Arial", color="D97706")
    ws.cell(sr+2, 1, "Total Absent").font  = Font(bold=True, name="Arial")
    ws.cell(sr+2, 2, absent).font          = Font(bold=True, name="Arial", color="DC2626")


def _safe_save(wb, path, tmp_path):
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        fallback = f"attendance_backup_{ts}.xlsx"
        wb.save(fallback)
        print(f"[!] {path} is open in Excel. Saved to {fallback} instead.")
    finally:
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except: pass


def save_to_excel(record):
    path = EXCEL_FILENAME; tmp_path = EXCEL_FILENAME + ".tmp"
    wb   = _load_or_create_wb(path)
    date_sheet = record["Date"]
    if date_sheet not in wb.sheetnames:
        ws = wb.create_sheet(date_sheet)
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 2:
            del wb["Sheet"]
        _write_header(ws)
    else:
        ws = wb[date_sheet]
    ws.append([record[h] for h in HEADERS])
    _style_row(ws, ws.max_row)
    _write_summary(ws)
    _safe_save(wb, path, tmp_path)


def rebuild_excel_for_today():
    today = datetime.now().strftime("%Y-%m-%d")
    path  = EXCEL_FILENAME; tmp_path = EXCEL_FILENAME + ".tmp"
    wb    = _load_or_create_wb(path)
    if today in wb.sheetnames: del wb[today]
    ws = wb.create_sheet(today)
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    _write_header(ws)
    with state_lock:
        records = [r for r in attendance_log if r["Date"] == today]
    for record in records:
        ws.append([record[h] for h in HEADERS])
        _style_row(ws, ws.max_row)
    _write_summary(ws)
    _safe_save(wb, path, tmp_path)


# ─────────────────────────────────────────────
#  ABSENT LIST GENERATOR
# ─────────────────────────────────────────────
def generate_absent_list():
    today = datetime.now().strftime("%Y-%m-%d")
    if not ROLL_LIST:
        return 0
    with state_lock:
        present_usns = {r["USN"] for r in attendance_log if r["Date"] == today}
    absent_usns = [u.strip().upper() for u in ROLL_LIST if u.strip().upper() not in present_usns]
    if not absent_usns:
        return 0
    path = EXCEL_FILENAME; tmp_path = EXCEL_FILENAME + ".tmp"
    wb   = _load_or_create_wb(path)
    if today not in wb.sheetnames:
        ws = wb.create_sheet(today)
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        _write_header(ws)
    else:
        ws = wb[today]
    for usn in absent_usns:
        record = {
            "USN": usn, "Name": "—", "Branch": "—", "Section": "—",
            "Date": today, "Time": "—", "Subject": SUBJECT, "Status": "Absent"
        }
        ws.append([record[h] for h in HEADERS])
        _style_row(ws, ws.max_row)
        with state_lock:
            attendance_log.append(record)
    _write_summary(ws)
    _safe_save(wb, path, tmp_path)
    return len(absent_usns)


# ─────────────────────────────────────────────
#  MONTHLY SUMMARY GENERATOR
# ─────────────────────────────────────────────
def generate_monthly_summary():
    path = EXCEL_FILENAME; tmp_path = EXCEL_FILENAME + ".tmp"
    if not os.path.exists(path):
        return False
    wb   = _load_or_create_wb(path)
    now  = datetime.now()
    month_prefix = f"{now.year}-{now.month:02d}-"
    date_sheets  = [s for s in wb.sheetnames if s.startswith(month_prefix)]
    if not date_sheets:
        return False
    all_usns = {}
    daily    = {}
    for sheet_name in date_sheets:
        ws  = wb[sheet_name]
        day = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            usn    = str(row[0])
            status = str(row[7]) if row[7] else "—"
            if status in ("Total Present", "Total Late", "Total Absent"): continue
            all_usns[usn] = {"Name": str(row[1] or "—"), "Branch": str(row[2] or "—"), "Section": str(row[3] or "—")}
            day[usn] = status
        daily[sheet_name] = day
    if "Monthly_Summary" in wb.sheetnames:
        del wb["Monthly_Summary"]
    ms = wb.create_sheet("Monthly_Summary", 0)
    active_dates = sorted(daily.keys())
    headers = ["USN", "Name", "Branch", "Section"] + active_dates + ["Present", "Late", "Absent", "Total Days", "Attendance %"]
    ms.append(headers)
    hdr_fill = PatternFill("solid", start_color="0F3460")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col_idx in range(1, len(headers) + 1):
        cell = ms.cell(row=1, column=col_idx)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = CENTER; cell.border = BORDER
        ms.column_dimensions[cell.column_letter].width = 14 if col_idx > 4 else 18
    pct_red    = PatternFill("solid", start_color="FEE2E2")
    pct_yellow = PatternFill("solid", start_color="FEF3C7")
    pct_green  = PatternFill("solid", start_color="D1FAE5")
    for usn, info in sorted(all_usns.items()):
        row_data = [usn, info["Name"], info["Branch"], info["Section"]]
        p = l = a = 0
        for d in active_dates:
            st = daily[d].get(usn, "Absent")
            row_data.append(st)
            if st == "Present": p += 1
            elif st == "Late":  l += 1
            else:               a += 1
        total = len(active_dates)
        pct   = round((p + l) / total * 100, 1) if total else 0
        row_data += [p, l, a, total, f"{pct}%"]
        ms.append(row_data)
        rn = ms.max_row
        for ci in range(1, len(row_data) + 1):
            cell = ms.cell(row=rn, column=ci)
            cell.font = CELL_FONT; cell.border = BORDER; cell.alignment = CENTER
            val = cell.value
            if val == "Present":  cell.fill = PRESENT_FILL
            elif val == "Late":   cell.fill = LATE_FILL
            elif val == "Absent": cell.fill = ABSENT_FILL
        pct_cell = ms.cell(row=rn, column=len(headers))
        if pct < 75:   pct_cell.fill = pct_red
        elif pct < 85: pct_cell.fill = pct_yellow
        else:          pct_cell.fill = pct_green
        pct_cell.font = Font(bold=True, name="Arial", size=10)
    ms.freeze_panes = "E2"
    _safe_save(wb, path, tmp_path)
    return True


# ─────────────────────────────────────────────
#  SETTINGS DIALOG
# ─────────────────────────────────────────────
class SettingsDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.result = None
        self.title("Settings")
        self.configure(bg="#1a1a2e")
        self.resizable(False, False)
        self.grab_set()
        self._build()

    def _build(self):
        tk.Label(self, text="Settings", font=("Segoe UI", 14, "bold"),
                 fg="white", bg="#1a1a2e").pack(pady=(16, 4))
        tk.Label(self, text="Changes apply immediately after Save",
                 font=("Segoe UI", 9), fg="#94a3b8", bg="#1a1a2e").pack(pady=(0, 12))
        frame = tk.Frame(self, bg="#1a1a2e")
        frame.pack(padx=20, pady=6)
        fields = [
            ("Subject Name",                     "subject",  SUBJECT),
            ("QR Rotate (seconds)",              "rotate",   str(QR_ROTATE_SECONDS)),
            ("Late After (minutes)",             "late",     str(LATE_AFTER_MINUTES)),
            ("Admin Password",                   "password", ADMIN_PASSWORD),
            ("Class Start Time (HH:MM or blank)","start",    CLASS_START_TIME or ""),
        ]
        self.vars = {}
        for label, key, default in fields:
            row = tk.Frame(frame, bg="#1a1a2e")
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label, font=("Segoe UI", 10), fg="#94a3b8",
                     bg="#1a1a2e", width=32, anchor="w").pack(side="left")
            var  = tk.StringVar(value=default)
            show = "*" if key == "password" else ""
            tk.Entry(row, textvariable=var, font=("Segoe UI", 11),
                     width=18, show=show).pack(side="left")
            self.vars[key] = var
        tk.Label(frame, text="Class Roll List (one USN per line):",
                 font=("Segoe UI", 10), fg="#94a3b8", bg="#1a1a2e",
                 anchor="w").pack(fill="x", pady=(10, 4))
        self.roll_text = tk.Text(frame, height=6, width=50,
                                 font=("Consolas", 9), bg="#0f172a", fg="#e2e8f0")
        self.roll_text.pack()
        self.roll_text.insert("end", "\n".join(ROLL_LIST))
        btn_frame = tk.Frame(self, bg="#1a1a2e", pady=14)
        btn_frame.pack()
        tk.Button(btn_frame, text="Save", command=self._save,
                  font=("Segoe UI", 10, "bold"), bg="#16a34a", fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left", padx=6)
        tk.Button(btn_frame, text="Cancel", command=self.destroy,
                  font=("Segoe UI", 10, "bold"), bg="#7f1d1d", fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left", padx=6)

    def _save(self):
        global SUBJECT, QR_ROTATE_SECONDS, LATE_AFTER_MINUTES, ADMIN_PASSWORD, CLASS_START_TIME, ROLL_LIST, class_start_dt
        try:
            rotate = int(self.vars["rotate"].get())
            late   = int(self.vars["late"].get())
        except ValueError:
            messagebox.showerror("Error", "Rotate and Late must be whole numbers.", parent=self)
            return
        subj  = self.vars["subject"].get().strip()
        pwd   = self.vars["password"].get().strip()
        start = self.vars["start"].get().strip()
        rolls = [u.strip().upper() for u in self.roll_text.get("1.0", "end").splitlines() if u.strip()]
        if not subj or not pwd:
            messagebox.showerror("Error", "Subject and Password cannot be empty.", parent=self)
            return
        SUBJECT = subj; QR_ROTATE_SECONDS = rotate; LATE_AFTER_MINUTES = late
        ADMIN_PASSWORD = pwd; CLASS_START_TIME = start if start else None; ROLL_LIST = rolls
        if CLASS_START_TIME:
            h, m = map(int, CLASS_START_TIME.split(":"))
            class_start_dt = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
        else:
            class_start_dt = datetime.now()
        self.result = True
        self.destroy()
        messagebox.showinfo("Saved", f"Settings saved!\nSubject: {SUBJECT}\nQR Rotate: {QR_ROTATE_SECONDS}s")


# ─────────────────────────────────────────────
#  TKINTER GUI
# ─────────────────────────────────────────────
class AttendanceApp(tk.Tk):
    def __init__(self, server_ip):
        super().__init__()
        self.server_ip = server_ip
        self.title("Dynamic QR Attendance System")
        self.configure(bg="#1a1a2e")
        self.resizable(False, False)
        self._build_ui()
        self._refresh_qr()
        self._tick()

    def _build_ui(self):
        hdr = tk.Frame(self, bg="#0f3460", pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Dynamic QR Attendance System",
                 font=("Segoe UI", 16, "bold"), fg="white", bg="#0f3460").pack()
        self.hdr_sub = tk.Label(hdr,
                 text=f"Subject: {SUBJECT}  |  {datetime.now().strftime('%A, %d %B %Y')}  |  Late after {LATE_AFTER_MINUTES} mins",
                 font=("Segoe UI", 10), fg="#94a3b8", bg="#0f3460")
        self.hdr_sub.pack(pady=(2, 0))

        body = tk.Frame(self, bg="#1a1a2e", padx=20, pady=16)
        body.pack()

        qr_frame = tk.Frame(body, bg="#ffffff", padx=10, pady=10,
                            highlightthickness=2, highlightbackground="#0f3460")
        qr_frame.grid(row=0, column=0, padx=(0, 20))
        self.qr_label = tk.Label(qr_frame, bg="white")
        self.qr_label.pack()

        right = tk.Frame(body, bg="#1a1a2e")
        right.grid(row=0, column=1, sticky="n")

        tk.Label(right, text="Next refresh in",
                 font=("Segoe UI", 11), fg="#94a3b8", bg="#1a1a2e").pack(anchor="w")
        self.countdown_var = tk.StringVar(value="30s")
        tk.Label(right, textvariable=self.countdown_var,
                 font=("Segoe UI", 36, "bold"), fg="#38bdf8", bg="#1a1a2e").pack(anchor="w")
        self.progress = ttk.Progressbar(right, length=200, mode="determinate",
                                        maximum=QR_ROTATE_SECONDS)
        self.progress.pack(anchor="w", pady=(4, 14))

        tk.Label(right, text="Scan URL",
                 font=("Segoe UI", 11), fg="#94a3b8", bg="#1a1a2e").pack(anchor="w")
        self.url_var = tk.StringVar()
        tk.Label(right, textvariable=self.url_var,
                 font=("Segoe UI", 9), fg="#7dd3fc", bg="#1a1a2e",
                 wraplength=210, justify="left").pack(anchor="w", pady=(0, 14))

        tk.Label(right, text="Present  |  Late",
                 font=("Segoe UI", 11), fg="#94a3b8", bg="#1a1a2e").pack(anchor="w")
        self.count_var = tk.StringVar(value="0  |  0")
        tk.Label(right, textvariable=self.count_var,
                 font=("Segoe UI", 28, "bold"), fg="#4ade80", bg="#1a1a2e").pack(anchor="w")

        tk.Label(right, text="Devices locked",
                 font=("Segoe UI", 11), fg="#94a3b8", bg="#1a1a2e").pack(anchor="w", pady=(10,0))
        self.device_var = tk.StringVar(value="0")
        tk.Label(right, textvariable=self.device_var,
                 font=("Segoe UI", 28, "bold"), fg="#a78bfa", bg="#1a1a2e").pack(anchor="w")

        btn_frame = tk.Frame(self, bg="#1a1a2e", pady=10)
        btn_frame.pack()
        for text, cmd, color in [
            ("Open Excel",     self._open_excel,      "#0f3460"),
            ("Admin Panel",    self._open_admin,      "#1e3a5f"),
            ("Absent List",    self._gen_absent,      "#713f12"),
            ("Monthly Report", self._gen_monthly,     "#164e63"),
            ("Reset Devices",  self._reset_devices,   "#312e81"),
            ("Settings",       self._open_settings,   "#1e3a5f"),
            ("Close",          self.destroy,           "#7f1d1d"),
        ]:
            tk.Button(btn_frame, text=text, command=cmd,
                      font=("Segoe UI", 8, "bold"),
                      bg=color, fg="white", relief="flat",
                      padx=8, pady=7, cursor="hand2").pack(side="left", padx=2)

        log_outer = tk.Frame(self, bg="#1a1a2e", padx=20, pady=0)
        log_outer.pack(fill="x", pady=(0, 16))
        tk.Label(log_outer, text="Recent check-ins",
                 font=("Segoe UI", 10), fg="#94a3b8", bg="#1a1a2e").pack(anchor="w")
        self.log_box = tk.Text(log_outer, height=5, width=72,
                               bg="#0f172a", fg="#e2e8f0",
                               font=("Consolas", 9), bd=0, state="disabled")
        self.log_box.pack(fill="x")

    def _make_qr_image(self, url):
        qr = qrcode.QRCode(version=2, error_correction=qrcode.constants.ERROR_CORRECT_H,
                           box_size=7, border=2)
        qr.add_data(url); qr.make(fit=True)
        return qr.make_image(fill_color="black", back_color="white")

    def _refresh_qr(self):
        with state_lock:
            token = current_token
        url = f"http://{self.server_ip}:{SERVER_PORT}/?token={token}"
        self.url_var.set(url)
        img   = self._make_qr_image(url)
        img   = img.resize((240, 240), Image.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        self.qr_label.configure(image=photo)
        self.qr_label._img = photo

    def _tick(self):
        with state_lock:
            remaining = max(0.0, token_expiry - time.time())
        self.countdown_var.set(f"{int(remaining) + 1}s")
        self.progress["value"] = remaining
        today = datetime.now().strftime("%Y-%m-%d")
        with state_lock:
            present      = sum(1 for r in attendance_log if r["Date"] == today and r["Status"] == "Present")
            late         = sum(1 for r in attendance_log if r["Date"] == today and r["Status"] == "Late")
            device_count = len(submitted_ips)
        self.count_var.set(f"{present}  |  {late}")
        self.device_var.set(str(device_count))
        if remaining < 0.3:
            self.after(400, self._refresh_qr)
        self._update_log()
        self.after(250, self._tick)

    def _update_log(self):
        with state_lock:
            recent = list(reversed(attendance_log[-5:]))
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        for r in recent:
            tag  = "[P]" if r["Status"] == "Present" else ("[L]" if r["Status"] == "Late" else "[A]")
            line = f"  {r['Time']}  |  {r['USN']:<14}  |  {r['Name']:<20}  |  {r['Branch']}-{r['Section']}  {tag}\n"
            self.log_box.insert("end", line)
        self.log_box.configure(state="disabled")

    def _open_excel(self):
        if os.path.exists(EXCEL_FILENAME):
            if os.name == "nt": os.startfile(EXCEL_FILENAME)
            else: os.system(f"xdg-open '{EXCEL_FILENAME}'")
        else:
            messagebox.showinfo("Info", "No attendance recorded yet.")

    def _open_admin(self):
        import webbrowser
        pwd = simpledialog.askstring("Admin Panel", "Enter admin password:", show="*")
        if pwd == ADMIN_PASSWORD:
            webbrowser.open(f"http://localhost:{SERVER_PORT}/admin?pwd={pwd}")
        elif pwd is not None:
            messagebox.showerror("Error", "Wrong password!")

    def _gen_absent(self):
        if not ROLL_LIST:
            messagebox.showinfo("Roll List Empty",
                "No roll list found.\nOpen Settings and add USNs under 'Class Roll List' first.")
            return
        if not messagebox.askyesno("Generate Absent List",
                "This will mark all students NOT in today's attendance as Absent.\nContinue?"):
            return
        count = generate_absent_list()
        if count == 0:
            messagebox.showinfo("Done", "All students have already marked attendance!")
        else:
            messagebox.showinfo("Done", f"{count} student(s) marked as Absent in Excel.")

    def _gen_monthly(self):
        ok = generate_monthly_summary()
        if ok:
            messagebox.showinfo("Done",
                "Monthly Summary sheet created!\n\nGreen=85%+  |  Yellow=75-84%  |  Red=Below 75%")
            if os.name == "nt": os.startfile(EXCEL_FILENAME)
        else:
            messagebox.showinfo("No Data", "No attendance data found for this month yet.")

    def _reset_devices(self):
        if not messagebox.askyesno("Reset Device Locks",
                "This will allow all devices to submit attendance again.\n"
                "Use this only if a student genuinely needs to resubmit.\nContinue?"):
            return
        with state_lock:
            submitted_ips.clear()
        messagebox.showinfo("Done", "Device locks cleared! All devices can submit again.")

    def _open_settings(self):
        dlg = SettingsDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self.hdr_sub.config(
                text=f"Subject: {SUBJECT}  |  {datetime.now().strftime('%A, %d %B %Y')}  |  Late after {LATE_AFTER_MINUTES} mins"
            )


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
def start_http_server():
    server = ThreadedHTTPServer(("0.0.0.0", SERVER_PORT), AttendanceHandler)
    server.serve_forever()


if __name__ == "__main__":
    ip = get_local_ip()
    if CLASS_START_TIME:
        h, m = map(int, CLASS_START_TIME.split(":"))
        class_start_dt = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
    else:
        class_start_dt = datetime.now()

    t1 = threading.Thread(target=rotate_token, daemon=True)
    t1.start()
    time.sleep(0.1)

    t2 = threading.Thread(target=start_http_server, daemon=True)
    t2.start()

    print(f"[OK] Server   : http://{ip}:{SERVER_PORT}")
    print(f"[OK] Admin    : http://{ip}:{SERVER_PORT}/admin?pwd={ADMIN_PASSWORD}")
    print(f"[OK] Subject  : {SUBJECT}")
    print(f"[OK] QR rotates every {QR_ROTATE_SECONDS}s")
    print(f"[OK] Late after {LATE_AFTER_MINUTES} mins")
    print(f"[OK] Excel    : {EXCEL_FILENAME}")

    app = AttendanceApp(ip)
    app.mainloop()
